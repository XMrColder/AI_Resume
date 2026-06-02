"""
Phase 7 - 导出 Excel
把批量筛选结果（batch.py 产出的 batch_results.json，或内存里的 report dict）
导出成带格式的 Excel：排名总表（含每维度得分）+ 岗位信息表，可只导 Top N。

命令行用法：
    python export_excel.py batch_results.json
    python export_excel.py batch_results.json --top 10 -o shortlist.xlsx

也提供 report_to_xlsx_bytes(report, top_n) 供 Streamlit 下载按钮直接用。
"""
import io
import json
import argparse
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
BASE_FONT = Font(name=FONT)
ALT_FILL = PatternFill("solid", fgColor="EEF1F8")


def _dim_score(dimensions, name):
    for d in dimensions:
        if d.get("name") == name:
            return d.get("score")
    return None


def build_workbook(report: dict, top_n: Optional[int] = None) -> Workbook:
    ranking = report.get("ranking", [])
    if top_n:
        ranking = ranking[:top_n]

    # 各候选人出现过的维度名 → 动态列（必备技能 / 经历相关 / ...）
    dim_names = []
    for r in ranking:
        for d in r.get("dimensions", []):
            if d.get("name") not in dim_names:
                dim_names.append(d["name"])

    wb = Workbook()
    ws = wb.active
    ws.title = "排名总表"

    headers = ["排名", "姓名", "匹配分", "文件", "匹配技能", "缺口"] + [f"{n}" for n in dim_names]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in ranking:
        row = [r.get("rank"), r.get("name") or "(未知)", r.get("score"), r.get("file"),
               "、".join(r.get("highlights", [])) or "—",
               "、".join(r.get("gaps", [])) or "—"]
        row += [_dim_score(r.get("dimensions", []), n) for n in dim_names]
        ws.append(row)

    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")
    num_cols = {1, 3} | set(range(7, 7 + len(dim_names)))
    wrap_cols = {5, 6}
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=1 + len(ranking))):
        for cell in row:
            cell.font = BASE_FONT
            if cell.column in num_cols:
                cell.alignment = center
            elif cell.column in wrap_cols:
                cell.alignment = wrap
            if i % 2 == 1:                       # 隔行浅底，便于阅读
                cell.fill = ALT_FILL

    widths = [6, 12, 8, 24, 36, 22] + [11] * len(dim_names)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 摘要：平均分用 Excel 公式（打开后自动重算）
    n = len(ranking)
    if n:
        srow = ws.max_row + 2
        ws.cell(srow, 2, "平均分").font = Font(name=FONT, bold=True)
        avg = ws.cell(srow, 3, f"=AVERAGE(C2:C{1 + n})")
        avg.font = Font(name=FONT, bold=True)
        avg.alignment = center

    # 第二个 sheet：岗位信息
    jd = report.get("jd", {})
    ws2 = wb.create_sheet("岗位信息")
    rows = [
        ("职位", jd.get("title") or "—"),
        ("最低年限", jd.get("min_years")),
        ("必备技能", "、".join(jd.get("required_skills", []))),
        ("加分项", "、".join(jd.get("preferred_skills", []))),
        ("岗位职责", "；".join(jd.get("responsibilities", []))),
    ]
    for label, value in rows:
        ws2.append([label, value])
    for row in ws2.iter_rows():
        row[0].font = Font(name=FONT, bold=True)
        row[1].font = BASE_FONT
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 64

    return wb


def export_to_excel(report: dict, path: str, top_n: Optional[int] = None):
    build_workbook(report, top_n).save(path)


def report_to_xlsx_bytes(report: dict, top_n: Optional[int] = None) -> bytes:
    buf = io.BytesIO()
    build_workbook(report, top_n).save(buf)
    return buf.getvalue()


def main():
    ap = argparse.ArgumentParser(description="Phase 7 导出 Excel")
    ap.add_argument("report_json", help="batch.py 产出的 batch_results.json")
    ap.add_argument("--top", type=int, default=None, help="只导出前 N 名")
    ap.add_argument("-o", "--out", default="screening_result.xlsx", help="输出文件名")
    args = ap.parse_args()

    with open(args.report_json, encoding="utf-8") as f:
        report = json.load(f)
    export_to_excel(report, args.out, args.top)
    n = len(report.get("ranking", []))
    n = min(n, args.top) if args.top else n
    print(f"已导出 {n} 名候选人到 {args.out}")


if __name__ == "__main__":
    main()
